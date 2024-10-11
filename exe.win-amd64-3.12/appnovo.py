import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import *
from selenium.webdriver.support.ui import WebDriverWait
from time import sleep
from openpyxl import load_workbook

def iniciar_driver():
    chrome_options = Options()
    arguments = ['--lang=en-US', '--window-size=1300,1000', '--incognito']
    for argument in arguments:
        chrome_options.add_argument(argument)

    chrome_options.add_experimental_option('prefs', {
        'download.prompt_for_download': False,
        'profile.default_content_setting_values.notifications': 2,
        'profile.default_content_setting_values.automatic_downloads': 1
    })

    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(
        driver,
        10,  # Tempo máximo de espera
        poll_frequency=1,
        ignored_exceptions=[
            NoSuchElementException,
            ElementNotVisibleException,
            ElementNotSelectableException,
        ]
    )
    return driver, wait

def buscar_emails(driver, wait):
    emails_encontrados = []
    try:
        body_text = driver.find_element(By.TAG_NAME, "body").text
        emails_encontrados = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", body_text)
        
        if emails_encontrados:
            print(f"E-mails encontrados: {emails_encontrados}")
        else:
            print("Nenhum e-mail encontrado.")
    except TimeoutException:
        print("Tempo esgotado: e-mail não encontrado.")
    except Exception as e:
        print(f"Erro ao buscar e-mail: {e}")
    
    return emails_encontrados

def limpar_prefixo_email(email):
    email = email.strip().replace('EMAIL', '').replace('email', '').strip()
    email_limpo = re.search(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", email)

    if email_limpo:
        return email_limpo.group(0)

    return None

def converter_tempo_pausa(pausa):
    """Converte a string de pausa escolhida pelo usuário em segundos."""
    if pausa.endswith("s"):
        return int(pausa[:-1])  # Remove 's' e converte para int
    elif pausa.endswith("min"):
        return int(pausa[:-3]) * 60  # Remove 'min' e converte para minutos
    elif pausa.isdigit():
        return int(pausa)  # Caso tenha sido passado um número sem sufixo
    return 0

def iniciar_selenium(arquivo_excel_path, url_quantity, tempo_pausa, update_pause_message):
    # Carregar o arquivo Excel selecionado
    arquivo_excel = load_workbook(arquivo_excel_path)
    sheet = arquivo_excel.active

    # Lista de URLs
    urls = []
    for linhas in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if linhas[0] is not None:
            url = linhas[0]
            urls.append(url)

    driver, wait = iniciar_driver()
    tempo_pausa = converter_tempo_pausa(tempo_pausa)

    total_urls = len(urls)  # Obter o total de URLs
    for start_idx in range(0, total_urls, url_quantity):
        batch_urls = urls[start_idx:start_idx + url_quantity]
        
        for idx, url in enumerate(batch_urls, start=start_idx + 2):
            try:
                print(f'Acessando o site: {url}')
                driver.get(url)
                sleep(3)

                # Buscar e-mails no site
                emails_encontrados = buscar_emails(driver, wait)

                if emails_encontrados:
                    emails_limpos = [limpar_prefixo_email(email) for email in emails_encontrados if limpar_prefixo_email(email)]
                    for col_idx, email in enumerate(emails_limpos, start=2):  # Começa na coluna B (2)
                        sheet.cell(row=idx, column=col_idx).value = email
                else:
                    sheet.cell(row=idx, column=2).value = 'Nenhum e-mail encontrado'

            except Exception as e:
                print(f"Erro ao acessar {url}: {e}")
                # Registra "site em erro" na coluna B (ou onde preferir)
                if 'invalid argument' in str(e).lower():
                    sheet.cell(row=idx, column=2).value = 'Site em erro'
                else:
                    sheet.cell(row=idx, column=2).value = 'Erro ao acessar site'

            sleep(3)

        print(f"Pausando por {tempo_pausa} segundos...")
        update_pause_message(f'execução pausada {tempo_pausa}(s) ')  # Atualiza a mensagem na interface
        sleep(tempo_pausa)
        update_pause_message("")  # Limpa a mensagem após a pausa
    
    arquivo_excel.save(arquivo_excel_path.replace(".xlsx", " Atualizado.xlsx"))
    driver.quit()

def interromper_automação(driver):
    driver.quit()